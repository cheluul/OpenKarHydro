# -*- coding: utf-8 -*-
"""
__author__ = "Liu dongdong Yangya"
__date__ = "2022-xx-xx"
__version__ = "1.0"
"""

import OpenKarHydro_properties
import openpyxl
import numpy as np
import pandas as pd
from tqdm import tqdm

class Laio(object):
    print("模型运行开始")
    def __init__(self,pre_input):
        self.goal = pre_input['goal']            
        self.Crop_type = pre_input['Crop_type']   
        self.parametsrs = pre_input['parameters']           
        self.order = pre_input['order']  
        OpenKarHydro_properties.Read_canshu(self.goal, self.Crop_type, self.order, self.parametsrs)
        self.para =  OpenKarHydro_properties.pro  
        #研究区地理位置
        self.Psi = self.para['Loc_thr']['Psi']     
        self.Z = self.para['Loc_thr']['Z']          
        self.ht = self.para['Loc_thr']['ht']        
        self.λ = self.para['Loc_thr']['λ']          
        #时间离散
        self.days = self.para['Time_Method']['days']    
        self.dt = self.para['Time_Method']['dt']       
        #气象数据
        self.J = self.para['Time_Method']['J']          
        self.T_max = self.para['Atmosphere']['T_max']
        self.T_mean = self.para['Atmosphere']['T_mean']
        self.T_min = self.para['Atmosphere']['T_min']
        self.Rh = self.para['Atmosphere']['Rh']         
        self.Rain = self.para['Atmosphere']['Rain']     
        self.U2 = self.para['Atmosphere']['U2']         
        self.h = self.para['Atmosphere']['h']           
        #实测分项数据
        self.s_actual = self.para['Actual_value']['s_actual']
        self.Irr = self.para['Actual_value']['Irr']     
        self.Cov = self.para['Actual_value']['Cov']/100 
        self.Roff = self.para['Actual_value']['Roff']   
        self.Kc = self.para['Actual_value']['Kc']      
        #土壤性质属性参数
        self.floor = np.arange(0,self.para['params'].shape[0],1) 
        self.ETw = self.para['params']['ETw']        
        self.Zr = self.para['params']['Zr']          
        self.Depth = self.para['params']['Depth']    
        self.Ks = self.para['params']['Ks']          
        self.n = self.para['params']['n']            
        self.sfc = self.para['params']['sfc']       
        self.st = self.para['params']['st']         
        self.sw = self.para['params']['sw']          
        self.sh = self.para['params']['sh']         
        self.Inter_max = self.para['params']['Inter_max']  
        self.beta = self.para['params']['beta']             
        self.s_init = self.para['params']['s_init']        
        self.grow_start = self.para['params']['Crop_start'] 
        self.grow_end = self.para['params']['Crop_end']    

    #时间离散
    def Time_Method(self):
        self.month = np.arange(0, (self.days + self.dt) / 30, self.dt) 
        self.day = np.arange(0, self.days + self.dt - 1, self.dt)       
        self.hours = np.arange(0, self.days * 24, self.dt)              

    #水分输入
    def water_input(self,tim):
        self.Rain[tim] = np.where(self.Rain[tim]>=50 and self.goal==3,50,self.Rain[tim])
        self.s_Input = self.Rain[tim] + self.Irr[tim]
        self.s_Input = np.where(self.s_Input > self.Ks[0], self.Ks[0], self.s_Input) 

    #水分输出
    def water_loss(self,tim,s):
        ##冠层截留
        self.Cov_max = np.max(self.Cov)
        if self.goal == 3:
            self.Cov[tim] = self.grow_Cov
            self.Kc[tim] = self.grow_Kc
            self.Cov_max = self.K
        self.Inter1 = self.Inter_max[0]*(self.Cov[tim]/self.Cov_max)
        self.Inter1 = np.where(0<self.Rain[tim]<=self.Inter1,self.Rain[tim],self.Inter1) if self.Rain[tim]!=0 else 0
        self.Inter2 = np.where(0<self.Rain[tim]<=17,0.55*self.Cov[tim]*self.Rain[tim]**(0.52-0.0085*(self.Rain[tim] - 5.0)),1.85*self.Cov[tim]) if self.Rain[tim]!=0 else 0

        #径流
        CN2 = 68 if self.Cov[tim] < 0.5 else 49 if 0.5 <= self.Cov[tim] <= 0.75 else 39    
        slp = 0.135  
        CN2 = CN2*(322.79+15.63*slp)/(slp+323.52) 
        CN1 = CN2-20*(100-CN2)/(100-CN2+np.exp(2.533-0.0636*(100-CN2)))
        CN3 = CN2*np.exp(0.00673*(100-CN2))
        Sum_Rain = np.sum([self.Rain[tim],self.Rain[tim-1],self.Rain[tim-2],self.Rain[tim-3],self.Rain[tim-4]]) 
        Crop_cond = self.grow_start[0]<=self.J[tim]<=self.grow_end[0]  
        CN = np.piecewise(Sum_Rain,
                        [(Crop_cond==True and Sum_Rain<35.6) or (Crop_cond==False and Sum_Rain<12.7),
                         (Crop_cond==True and (35.6<=Sum_Rain<=53.3)) or (Crop_cond==False and (12.7<=Sum_Rain<=27.9)),
                         (Crop_cond==True and Sum_Rain>53.3) or (Crop_cond==False and Sum_Rain<27.9)],
                          [CN1,CN2,CN3])
        S = 25400/CN-254 
        self.Ia = float(self.λ) * S  
        self.Roff = (self.Rain[tim]-self.Ia)**2/(self.Rain[tim]+S-self.Ia) if self.Rain[tim] >= self.Ia else 0

        #蒸散发
        α = 0.23    
        a = 0.25    
        b = 0.50    
        G = 0       
        gsc = 0.0820 
        sigma = 4.903 * 10 ** (-9) 
        P = 101.3 * ((293 - 0.0065 * self.Z) / 293) ** 5.26  
        Gama = 0.665 * 10 ** (-3) * P  
        es_max = 0.6108 * np.exp(17.27 * self.T_max / (self.T_max + 237.3))  
        es_min = 0.6108 * np.exp(17.27 * self.T_min / (self.T_min + 237.3)) 
        es = (es_max + es_min) / 2       
        ea = np.mean(self.Rh)/100 * es  
        Delta = 0.409 * np.sin(2 * np.pi / 365 * self.J - 1.39) 
        Ws = np.arccos(-np.tan(self.Psi) * np.tan(Delta))       
        H = 24 / np.pi * Ws    
        dr = 1 + 0.033 * np.cos(2 * np.pi / 365 * self.J) 
        Ra = 24 * 60 / np.pi * gsc * dr * (Ws * np.sin(self.Psi) * np.sin(Delta) + np.cos(self.Psi) * np.cos(Delta) * np.sin(Ws))   
        Rs = (a + b * self.h / H) * Ra  
        Rso = (a + b) * Ra              
        Rn1 = sigma * ((self.T_max + 272.15) ** 4 + (self.T_min + 272.15) ** 4) / 2 * (0.34 - 0.14 * (ea ** 0.5)) * (1.35 * Rs / Rso - 0.35) 
        Rns = (1 - α) * Rs   
        Rn = Rns - Rn1       
        Deta = 4098 * (0.6108 * np.exp(17.27 * self.T_mean / (self.T_mean + 273.3))) / (self.T_mean + 273.3) ** 2  
        self.pm_fao56 = (0.408 * Deta * (Rn - G) + (Gama * 900 * self.U2 * (es - ea)) / (self.T_mean + 273)) / (Deta + Gama * (1 + 0.34 * self.U2))  

        self.AET,self.Lw = [],[]
        for floor in self.floor[0:]:
            if self.Depth[0] <= self.ht <= self.Depth[1]:
                self.Kh = 1 if floor == 0 else 0
            if self.Depth[1] < self.ht <= self.Depth[2]:
                self.Kh = 1 if floor == 0 else 0.048*(self.Depth[floor]/1000)**(-0.048) if floor == 1 else 0
            else:
                self.Kh = 1 if floor == 0 else 0.048*(self.Depth[floor]/1000)**(-0.048)
        
            self.ETmax = self.pm_fao56[tim] * self.Kc[tim] * self.Kh
            self.AET_Laio = np.piecewise(s[floor],[s[floor] <= self.sh[floor],(s[floor] <= self.sw[floor]) & (s[floor] > self.sh[floor]),
                                         (s[floor] <= self.st[floor]) & (s[floor] > self.sw[floor]),s[floor] > self.st[floor]],
                                         [0,self.ETw[floor] * ((s[floor] - self.sh[floor]) / (self.sw[floor] - self.sh[floor])),
                                          self.ETw[floor] + (self.ETmax - self.ETw[floor]) * ((s[floor] - self.sw[floor]) / (self.st[floor] - self.sw[floor])),
                                          self.ETmax])
            self.AET_Laio = np.where(self.Rain[tim]==0,0,self.AET_Laio) 

            #渗漏
            self.Lw1 = np.piecewise(s[floor],[s[floor] < self.sfc[floor],s[floor]>= self.sfc[floor]],
                                   [0,self.Ks[floor] * (np.exp(self.beta[floor] * (s[floor] - self.sfc[floor])) - 1) / (np.exp(self.beta[floor] * (1 - self.sfc[floor])) - 1)])
            self.AET.append(float(self.AET_Laio))
            self.Lw.append(float(self.Lw1))

    #多层净含水率
    def Net_s(self):
        self.Net_s0 = ((self.s_Input + self.AET[1]) - (self.Inter1 + self.AET[0] + self.Lw[0]+self.Roff))/(self.n[0] * self.Zr[0])
        self.Net_s1 = ((self.Lw[0] + self.AET[2]) - (self.AET[1] + self.Lw[1])) / (self.n[1] * self.Zr[1])
        self.Net_s2 = (self.Lw[1] - self.AET[2] - self.Lw[2]) / (self.n[2] * self.Zr[2])
        self.rhs = lambda Data: np.array([self.Net_s0,self.Net_s1,self.Net_s2])

    #解常微分方程
    def euler_forward(self, Data,dt):
        return Data + dt * self.rhs(Data)

    def improved_euler(self, Data,dt):
        yp = Data + dt * self.rhs(Data)
        return Data + 0.5 * dt * (self.rhs(Data) + self.rhs(yp))

    def Runge_Kutta4(self, Data, dt):
        k1 = dt * self.rhs(Data)
        k2 = dt * self.rhs(Data + 0.5 * k1)
        k3 = dt * self.rhs(Data + 0.5 * k2)
        k4 = dt * self.rhs(Data + k3)
        return np.array(Data + (k1 + 2.0 * (k2 + k3) + k4) / 6.0)

    def run_Laio_RK4(self):
        T, self.Sim = self.days, []
        s = np.array(self.s_init)  
        Cov0 = 0.25  
        results_avg = np.array([[0, 0, 0, 0, 0, 0, 0, 0, 0]])  
        for tim in tqdm(np.arange(0,T,1)):  
            self.water_input(tim)
            if self.goal == 3: 
                self.Crop_growth(Cov0,s[0],tim)
            self.water_loss(tim, s)
            self.Net_s()
            self.Sim.append(s)

            iters = 0  
            re_error = 0.0001  
            w = np.array(s * self.n * self.Zr)  
            Cov = self.Cov[tim] if self.goal == 1 or 2 else self.grow_Cov
            ends = np.array([[tim, s.tolist(), w.tolist(),self.Rain[tim], self.AET, self.Inter1, self.Lw, self.Roff, Cov]])
            results_avg = np.concatenate((results_avg, ends))
            while True:  
                h = 0.0001 
                iters = iters + 1
                s1 = self.Runge_Kutta4(s,h)   
                s2 = self.Runge_Kutta4(s1,h)  
                if (abs((s2-s1)/s1)< re_error).all() == True:
                     print('迭代收敛跳出本循环，此时的迭代次数为{}'.format(iters))
                     break
                else:
                    print("第{}次迭代值不满足收敛要求，继续迭代".format(iters))
                    s = s2
                    continue
            s = self.Runge_Kutta4(s2, self.dt) 
            s = np.where(s>1,1,s)
            Cov0 = self.grow_Cov if self.goal== 3 else None          
        results_avg = pd.DataFrame(results_avg,
                                   columns=['tim(day)',
                                            'Volumetric soil water content (nondim)',
                                            'Water storage(mm)',
                                            'Rain (mm/d)',
                                            'AET_imitate (mm/d)',
                                            'Inter_imitate (mm/d)',
                                            'Lw_imitate (mm/d)',
                                            'Roff_imitate (mm/d)',
                                            'Grow_Cov'])
        results_avg = results_avg[1:]  
        
        #模型反演
        if self.goal == 1:
            results_avg.to_csv('result/MC_result/Simulated_fenxiang/Crop{}/第{}次反演模拟.txt'.format(self.Crop_type, self.order),index=False)
            wb = openpyxl.load_workbook("result/MC_result/Simulated_s/Crop{}_MC_s.xlsx".format(self.Crop_type))
            for day, S_day in enumerate(self.Sim, 1):
                for floor, name in enumerate(wb.sheetnames):
                    ws = wb[name]
                    ws.cell(row=day, column=self.order).value = S_day[floor]
                    wb.save("result/MC_result/Simulated_s/Crop{}_MC_s.xlsx".format(self.Crop_type))
            wb.close()  
            
        #模型验证
        if self.goal == 2:
            results_avg.to_csv('result/ME_result/Simulated_fenxiang/Crop{}/第{}次正演验证.txt'.format(self.Crop_type, self.order), index=False)
            wb = openpyxl.load_workbook("result/ME_result/Simulated_s/Crop{}_ME_s.xlsx".format(self.Crop_type))
            for day, S_day in enumerate(self.Sim, 1):
                for floor, name in enumerate(wb.sheetnames):
                    ws = wb[name]
                    ws.cell(row=day, column=self.order).value = S_day[floor]
                    wb.save("result/ME_result/Simulated_s/Crop{}_ME_s.xlsx".format(self.Crop_type))
            wb.close() 

        #敏感性分析
        if self.goal == 4:
            fy_t = [0, 16, 31, 45, 61, 68, 81, 94, 116, 127, 137, 159, 181]
            self.Sim_s0, self.Sim_s1, self.Sim_s2 = [], [], []
            for s, index in zip(self.Sim, np.arange(len(self.Sim))):
                if index in fy_t:
                    self.Sim_s0.append(s[0]), self.Sim_s1.append(s[1]), self.Sim_s2.append(s[2])
            self.Sim_s0, self.Sim_s1, self.Sim_s2 = np.array(self.Sim_s0), np.array(self.Sim_s1), np.array(self.Sim_s2)
            Sensitivity_data = np.hstack((self.Sim_s0, self.Sim_s1, self.Sim_s2))
            return Sensitivity_data
        
        #反演与验证
        if self.goal == 1 or 2:
            fy_t = [0, 32, 43, 74, 108, 145, 166, 191] if self.goal == 1 else [0, 16, 31, 45, 61, 68, 81, 94, 116, 127, 137, 159, 181]
            self.Sim_s0, self.Sim_s1, self.Sim_s2 = [], [], []
            self.Obs_s0, self.Obs_s1, self.Obs_s2 = self.s_actual[0], self.s_actual[1], self.s_actual[2]
            for s, index in zip(self.Sim, np.arange(len(self.Sim))):
                if index in fy_t:
                    self.Sim_s0.append(s[0]), self.Sim_s1.append(s[1]), self.Sim_s2.append(s[2])
            self.Sim_s0, self.Sim_s1, self.Sim_s2 = np.array(self.Sim_s0), np.array(self.Sim_s1), np.array(self.Sim_s2)
            fy_data = ([self.Sim_s0, self.Obs_s0], [self.Sim_s1, self.Obs_s1], [self.Sim_s2, self.Obs_s2])
            return fy_data

        #情景分析
        if self.goal == 3:
            results_avg.to_csv('result/SA_result/Crop{}/第{}次未来30年土壤水分变化.txt'.format(self.Crop_type, self.order),index=False)
    #生长计算模块(自然撂荒)
    def Crop_growth(self,Cov0,s0,tim):
        if self.Crop_type == 1:
            self.grow_Cov = self.Cov[tim]
            self.grow_Kc = self.Kc[tim]
            self.K = np.max(self.Cov)
        elif self.Crop_type == 2 or 3:
            self.Dm, self.P = self.sfc[0]*self.n[0], 2
            [self.K,self.r,self.L] = [0.8,5.446*0.001,5.196*0.001] if self.Crop_type == 2 else [0.85,5.479*0.001,2.632*0.001]
            self.R = self.r * Cov0 * (1 - Cov0 / self.K) * self.dt  
            self.D_R = self.L * (s0*self.n[0] - self.Dm) ** self.P / (((s0*self.n[0] - self.Dm) ** self.P)+self.sh[0]*self.n[0]**self.P)*self.dt  
            self.grow_Cov = Cov0 + (self.R - self.D_R)
            self.grow_Kc = np.where(self.grow_Cov > 0.25,1.59274*self.grow_Cov-0.03876,0.3) if self.Crop_type == 2 else np.where(self.grow_Cov > 0.05,0.94195*self.grow_Cov+0.43851, 0.45)


