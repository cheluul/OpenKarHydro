# !/usr/bin/env python
# -*- coding:utf-8 -*-
#Scenario Analysis
#对1990-2020年的气象数据内的各气象因子以年为单位进行随机组合，拟作2021-2051年的数据，对齐进行情景分析
#随机获取的未来气象数据年内和年际变异特征与1990-2020年观测的数据相一致。本研究不考虑未来气候变化情景。
#导入所需要的包
import pandas as pd
import numpy as np
import os
import openpyxl
import random

year_list = np.linspace(2021,2050,30) #生成1990-2020年的年列表
file_path = r"data/SA_data"
file_name = "shenmu 2021-2050.xlsx"
file_result = os.path.join(file_path, file_name)
df = pd.read_excel(r"data/SA_data/shenmu 2021-2050.xlsx")

def excel_spilt():
    wb = openpyxl.load_workbook(file_result)  #首先获取excel文件生成对象wb 使用openpyxl.load_workbook 方法
    index = 0  #设置一个数字变量
    for i in year_list:  #循环31次，31次来源于year_list列表长度
        count = 1  #定义一个数字变量
        sh1 = wb.create_sheet(str(year_list[index]))  #使用使用wb.create 方法创建工作薄名称。名称为列表中的值，index是上面定义的数字变量
        for rows in df['序号']:  #循环工作薄表中的内容
            if df['日期'][rows].year == i:
                sh1["A1"] = "日期"
                sh1["B1"] = "时间序列"
                sh1["C1"] = "最高气温"
                sh1["D1"] = "最低气温"
                sh1["E1"] = "平均气温"
                sh1["F1"] = "平均相对湿度"
                sh1["G1"] = "2m平均风速"
                sh1["H1"] = "实际日照时数"
                sh1["I1"] = "降雨量"
                sh1["J1"] = "灌溉"
                sh1["K1"] = "径流"
                sh1["A" + str(count+1)] = df['日期'][rows]  # A +　str(count) 第二行开始 写入
                sh1["B" + str(count+1)] = df['时间序列'][rows]
                sh1["C" + str(count+1)] = df['最高气温'][rows]
                sh1["D" + str(count+1)] = df['最低气温'][rows]
                sh1["E" + str(count+1)] = df['平均气温'][rows]
                sh1["F" + str(count+1)] = df['平均相对湿度'][rows]
                sh1["G" + str(count+1)] = df['2m平均风速'][rows]
                sh1["H" + str(count+1)] = df['实际日照时数'][rows]
                sh1["I" + str(count+1)] = df['降雨量'][rows]
                sh1["J" + str(count+1)] = df['灌溉'][rows]
                sh1["K" + str(count+1)] = df['径流'][rows]
                count += 1  #每完成一次循环 count + 1
        index += 1  #同时index + 1
    wb.save("data/SA_data/数据已按年份拆分到新的工作簿.xls")  #当函数全部执行完成后，使用wb.save保存

def N_random(N_sum,Crop_type):
    excel_spilt()
    #三种作物都分三种情景进行分析
    n = 0
    while n < N_sum:  #生成随机气象数据N次。
        #从31年中随机读取30年的气象数据作为2021-2050共30年的气象数据
        def random_data():
            years_data = pd.read_excel('data/SA_data/数据已按年份拆分到新的工作簿.xls', sheet_name=None)  ##获取目录下所有Excel中sheet
            del years_data["Sheet1"]
            J, Tmax, Tmin, Tmean, Rh, U2, h, Rain, Irr,  Roff, Cov, Kc = [], [], [], [], [], [], [], [], [], [], [], []
            years = np.linspace(2021, 2050, 30)
            Weather_data = random.sample(list(years_data.values()), 30)  ##从1990-2020年终随机生成以年为单位的气象数据--》2021-2050
            new_data = dict(zip(years, Weather_data))
            for key, value in new_data.items():
                J.extend(value["时间序列"])
                Tmax.extend(value['最高气温'])
                Tmin.extend(value['最低气温'])
                Tmean.extend(value['平均气温'])
                Rh.extend(value['平均相对湿度'])
                U2.extend(value['2m平均风速'])
                h.extend(value['实际日照时数'])
                Rain.extend(value['降雨量'])
                Irr.extend(value['灌溉'])
                Roff.extend(value['径流'])

            data = {'年日序数':J,'最高温度': Tmax, '最低温度': Tmean, '平均温度': Tmean,'湿度': Rh,
                    '风速': U2, '实际日照时数': h, '降雨': Rain,'灌溉': Irr, '径流': Roff}
            if len(Tmax) > len(Kc):
                ends = '1/1/2051'
            if len(Tmax) <= len(Kc):
                ends = '12/31/2050'
            dates = pd.date_range(start='1/1/2021', end=ends, name='时间')
            New_data = pd.DataFrame(data, index=dates,dtype=np.float)
            New_data = New_data.dropna(axis=0, how='any')
            New_data.to_excel('data/SA_data/未来30年气象预测/第{}次未来30年气象数据.xls'.format(n))
            print("第{}次生成随机气象数据".format(n))
        n += 1
        random_data()
N_random(1000,1)








