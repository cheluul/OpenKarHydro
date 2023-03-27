# -*- coding: utf-8 -*-
"""
Created on Wed March 10 2021
Class used to run the system. Includes integration of ds/dt term, includes the calculation of related sub-items of the Laio model.
This class runs a simple soil water balance model in order to simulate the soil moisture, evapotranspiration, leakage, andrunoff, based on daily rainfall data and some soil parameters.
__author__ = Yangya
__email__ = 3100862389@qq.com
__date__ = 2021-05-01
__version__ = 1.0.0

一、分项模块参数集
1.Time_Method:时间离散方法
2.Soil_layer:土壤层
3.ET_Method:蒸发损失方法
3.Lw_Method:渗漏损失方法
4.Inter_Method:冠层截留损失方法
二、分项模块计算
1.时间离散（月、日、小时）
2.水分输入计算（降雨，灌溉，入流）
3.水分输出计算（蒸发，冠层截留，渗漏，出流）
4.体积含水率计算s（t）
二、编程日志
2021.3.10:搭建基本框架
2021.3.11:完成时间离散与水分输入计算
2021.3.12-3.14:完成水分输出计算
2021.3.15:补充输入输出计算方法
2021.3.16:精简相关计算方法参数，创建problem字典作为参数集
2021.3.20:补充土壤含水率分层计算
2021.3.21-3.23:创建执行文件scrip
2021.3.24-3.26:进行程序debug
2021.3.27-3.30:添加图表及理顺层与层之间的连接
2021.4.01~4.：数据验证
"""

import Laio_properties
import openpyxl
import numpy as np
import pandas as pd
import pyet as pyet
import math
from tqdm import tqdm

Model = "CSIRO_MK_3.6.0"
class Laio(object):
    print("模型运行开始")
    def __init__(self,pre_input):
        #继承属性参数集
        self.goal = pre_input['goal']             #模型目的【校准、验证、情景分析】
        self.X = pre_input['X']                   #土地利用转换前植被A种植的年限
        self.Crop_typeA = pre_input['Crop_typeA']
        self.Crop_type = pre_input['Crop_typeB']  
        self.parametsrs = pre_input['parameters'] #七个校准参数
        self.order = pre_input['order']  #goal==1/2:模型（正演、反演）验证时第order次帕累托排序为1;goal==3:第order次未来气象数据/气候变化类型
        Laio_properties.Read_canshu(self.goal, self.X, Model, self.Crop_type, self.order, self.parametsrs)
        self.para = Laio_properties.pro  #继承数据前处理
        #研究区地理位置状况
        self.Psi = self.para['Loc_thr']['Psi']      #纬度弧度数[ψ]【rad】
        self.Z = self.para['Loc_thr']['Z']          #海拔【m】
        self.ht = self.para['Loc_thr']['ht']  #最大潜在蒸发深度
        self.λ = self.para['Loc_thr']['λ']            #初损系数
        #与时间离散相关的参数
        self.days = self.para['Time_Method']['days']    #计算时长【d】
        self.dt = self.para['Time_Method']['dt']        #时间步长
        #气象数据
        self.J = self.para['Time_Method']['J']          #日序数【-】
        self.T_max = self.para['Atmosphere']['T_max']     #最高气温【℃】*修正因子
        self.T_mean = self.para['Atmosphere']['T_mean']  #平均气温【℃】*修正因字
        self.T_min = self.para['Atmosphere']['T_min']    #最低气温【℃】*修正因子
        self.Rh = self.para['Atmosphere']['Rh']         #相对湿度【%】
        if self.goal == 3:
            Rain_factor = 0.45 if Model == "Can_ESM2" else 0.64 if Model == "CSIRO_MK_3.6.0" else 1.535
        else:
            Rain_factor = 1
        self.Rain = self.para['Atmosphere']['Rain']*Rain_factor    #降雨量【mm/d】*修正因子【0.45,0.64,1.535】
        self.U2 = self.para['Atmosphere']['U2']         #2m高风速【m/s】
        self.h = self.para['Atmosphere']['h']           #实际日照数【h/d】
        #实测分项数据
        self.s_actual = self.para['Actual_value']['s_actual'] #实测含水率【cm3*cm-3】
        self.Irr = self.para['Actual_value']['Irr']     #灌溉量【mm/d】
        self.Cov = self.para['Actual_value']['Cov']/100 #作物盖度【0-1】
        self.Roff = self.para['Actual_value']['Roff']   #径流量【mm/d】
        self.Kc = self.para['Actual_value']['Kc']       #作物系数
        #土壤性质属性参数
        self.floor = np.arange(0,self.para['params'].shape[0],1)  #根据输入土壤参数的形状（几行）--》判断土壤层数
        self.ETw = self.para['params']['ETw']        #植物萎焉临界的蒸发【mm/d】
        self.Zr = self.para['params']['Zr']          #模拟层土壤厚度【mm】
        self.Depth = self.para['params']['Depth']    #土层上边界距离大气边界的深度【mm】
        self.Ks = self.para['params']['Ks']          #饱和导水率【mm/d】
        self.n = self.para['params']['n']            #土壤孔隙度【-】
        self.sfc = self.para['params']['sfc']        #田间持水率【cm3*cm-3】
        self.st = self.para['params']['st']          #临界含水率阈值,开始收到水分胁迫[s*]【cm3*cm-3】
        self.sw = self.para['params']['sw']          #永久枯萎点土壤含水率【cm3*cm-3】
        self.sh = self.para['params']['sh']          #吸湿点土壤含水率【cm3*cm-3】
        self.Inter_max = self.para['params']['Inter_max']   #最大截留量【mm/d】
        self.beta = self.para['params']['beta']             #计算渗漏相关的空隙分布参数[β]【-】
        self.s_init = self.para['params']['s_init']         #初始含水率
        self.grow_start = self.para['params']['Crop_start'] #作物生长期开始日序数
        self.grow_end = self.para['params']['Crop_end']     #作物生长期结束日序数

    ##时间离散
    def Time_Method(self):
        self.month = np.arange(0, (self.days + self.dt) / 30, self.dt)  #按月离散
        self.day = np.arange(0, self.days + self.dt - 1, self.dt)       #按天离散
        self.hours = np.arange(0, self.days * 24, self.dt)              #按小时离散

    ##水分输入
    def water_input(self,tim):
        self.s_Input = self.Rain[tim] + self.Irr[tim]
        self.s_Input = np.where(self.s_Input > self.Ks[0], self.Ks[0], self.s_Input) #输入速率与饱和导水率比较，小于饱和导水率按输入速率输入，大于饱和导水率按饱和导水率输入
        
    ##水分输出
    def water_loss(self,tim,s):
        ##冠层截留计算
        self.Cov_max = np.max(self.Cov)  
        if self.goal == 3:
            self.Cov[tim] = self.grow_Cov
            self.Kc[tim] = self.grow_Kc
            self.Cov_max = self.K
        self.Inter1 = self.Inter_max[0]*(self.Cov[tim]/self.Cov_max)
        self.Inter1 = np.where(0<self.Rain[tim]<=self.Inter1,self.Rain[tim],self.Inter1) if self.Rain[tim] !=0 else 0   #假设为作物盖度与最大截留阈值的函数
        self.Inter2 = np.where(0<self.Rain[tim]<=17,0.55*self.Cov[tim]*self.Rain[tim]**(0.52-0.0085*(self.Rain[tim] - 5.0)),1.85*self.Cov[tim]) if self.Rain[tim]!=0 else 0
    
        ##径流计算:SCS-CN模型  
        CN2 = 68 if self.Cov[tim] < 0.5 else 49 if 0.5 <= self.Cov[tim] <= 0.75 else 39    #反应前期集水区特性的参数（0<=CN<=100）
        slp = 0.135  #田间平均坡度
        CN2 = CN2*(322.79+15.63*slp)/(slp+323.52) #根据坡度改进的CN2
        CN1 = CN2-20*(100-CN2)/(100-CN2+np.exp(2.533-0.0636*(100-CN2)))  #计算CN1
        CN3 = CN2*np.exp(0.00673*(100-CN2))  #计算CN3
        Sum_Rain = np.sum([self.Rain[tim],self.Rain[tim-1],self.Rain[tim-2],self.Rain[tim-3],self.Rain[tim-4]])  #前五天降雨状况
        Crop_cond = self.grow_start[0]<=self.J[tim]<=self.grow_end[0]  #判断作物状态【生长期/休眠期】
        CN = np.piecewise(Sum_Rain,
                        [(Crop_cond==True and Sum_Rain<35.6) or (Crop_cond==False and Sum_Rain<12.7),
                         (Crop_cond==True and (35.6<=Sum_Rain<=53.3)) or (Crop_cond==False and (12.7<=Sum_Rain<=27.9)),
                         (Crop_cond==True and Sum_Rain>53.3) or (Crop_cond==False and Sum_Rain<27.9)],
                          [CN1,CN2,CN3])  
        S = 25400/CN-254 #潜在最大降雨损失量【mm】
        self.Ia = float(self.λ) * S  #Ia:集水区初损量, λ:初损系数（0<=λ<=0.4）
        self.Roff = (self.Rain[tim]-self.Ia)**2/(self.Rain[tim]+S-self.Ia) if self.Rain[tim] >= self.Ia else 0     
        
        #日尺度侵蚀量计算：MUSLE模型【A=11.8*(QqArea)**0.56*Ke*LS*C*P】    
        q_peak = self.Roff*60/(1000*24*60*60)  #洪峰径流【m3/s】
        Area = 15*4/10000 #【hm2】
        Cl,Si,Sa,Sn,SOC = [7.136,54.451,38.413,0.616,0.216] if self.Crop_type==1 else [15.193,57.927,26.870,0.713,0.225] if self.Crop_type==2 else [7.886,56.648,35.466,0.645,0.252]
        self.K_RUSLE = (0.2+0.3*np.exp(-0.0256*Sa*(1-Si/100)))*(Si/(Cl+Si))**0.3*(1-0.25*SOC/(SOC+np.exp(3.72-2.95*SOC)))*(1-0.7*Sn/(Sn+np.exp(-5.51+22.9*Sn))) #Sharplyet al.,(1990)
        slope,λ,m = 13.5,15,0.5 #坡度;坡长;坡长指数(0.5 for slopes > 9% but < 48% as suggested by McCool et al. (1993).
        self.L = (λ/22.3)**m   #采用刘宝元在CSLE中改进的坡长因子   
        self.S = 10.80*np.sin(slope)+0.03 if self.Crop_type== 1 else 16.80*np.sin(slope)-0.5 if self.Crop_type==2 else 21.91*np.sin(slope)-0.96 #对于坡度<10°的采用McCool et al.,1987,>10°的采用Liu et al.,1994
        self.LS_RUSLE = self.L*self.S
        self.C_RUSLE = 1 if self.Cov[tim]==0 else 0.6508-0.3436*math.log(self.Cov[tim],math.e) if 0 < self.Cov[tim] <= 0.783 else 0#Crop and soil management
        self.C_RUSLE = np.where(self.C_RUSLE>1,1,self.C_RUSLE)
        self.P_RUSLE = 0.3 if self.Crop_type == 1 else 0.16 
        rock = 0  #rock砾石含量
        CFRG = np.exp(-0.053*rock) #粗碎屑因子
        self.A = 11.8*(self.Roff/Area*q_peak*Area)**0.56*self.K_RUSLE*self.LS_RUSLE*self.C_RUSLE*self.P_RUSLE*CFRG #【t】
        self.A = self.A *1000000 #因径流小区面积较小，为了比较效果t-->g
        
        ##根据[allen_1998]--pm_fao56计算潜在蒸发量[mm/d]
        G = 0        #大地热通量【MJ /(m2.d)】
        P = 101.3 * ((293 - 0.0065 * self.Z) / 293) ** 5.26  #大气压【kpa】
        Gama = 0.665 * 10 ** (-3) * P  # 温度表常数【Kpa/℃】[γ]
        es_max = 0.6108 * np.exp(17.27 * self.T_max / (self.T_max + 237.3))  #按每日最高温度计算的饱和水汽压【Kpa】
        es_min = 0.6108 * np.exp(17.27 * self.T_min / (self.T_min + 237.3))  #按每日最低温度计算的饱和水汽压【Kpa】
        es = (es_max + es_min) / 2       #平均饱和水汽压【Kpa】
        ea = np.mean(self.Rh)/100 * es   #实际水汽压【Kpa】
        Deta = 4098 * (0.6108 * np.exp(17.27 * self.T_mean / (self.T_mean + 273.3))) / (self.T_mean + 273.3) ** 2  # 在空气温度T时的饱和水汽压斜率[Δ]【Kpa/℃】
        Rn = self.h
        self.pm_fao56 = (0.408 * Deta * (Rn - G) + (Gama * 900 * self.U2 * (es - ea)) / (self.T_mean + 273)) / (Deta + Gama * (1 + 0.34 * self.U2))  #彭曼公式计算的潜在蒸发【mm/d】
        
        self.AET,self.Lw = [],[]
        for floor in self.floor[0:]:
            #1：设定最大蒸发深度【ht】，表层的蒸发能力为1，最大潜在蒸发深度处的蒸发能力为0。
            #2：拟定Kh表示各层上边界的蒸发能力，Kc表示作物系数，各层的最大蒸发表示为Emax（s，h，t）= PET*Kh*Kc
            #3：再带入Laio模型的蒸发公式【考虑考虑胁迫的影响】得到最终的实际蒸发
            #A：幂指数函数,浅根作物不考虑深层蒸发
            if self.Depth[0] <= self.ht <= self.Depth[1]:
                self.Kh = 1 if floor == 0 else 0
            elif self.Depth[1] < self.ht <= self.Depth[2]:
                self.Kh = 1 if floor == 0 else 0.048*(self.Depth[floor]/1000)**(-0.048) if floor == 1 else 0
            else:
                self.Kh = 1 if floor == 0 else 0.048*(self.Depth[floor]/1000)**(-0.048)
            #Laio模型蒸发计算【四阶段】
            self.ETmax = self.pm_fao56[tim]*self.Kc[tim] * self.Kh
            self.AET_Laio = np.piecewise(s[floor],[s[floor] <= self.sh[floor],(s[floor] <= self.sw[floor]) & (s[floor] > self.sh[floor]),
                                         (s[floor] <= self.st[floor]) & (s[floor] > self.sw[floor]),s[floor] > self.st[floor]],
                                         [0,self.ETw[floor] * ((s[floor] - self.sh[floor]) / (self.sw[floor] - self.sh[floor])),
                                          self.ETw[floor] + (self.ETmax - self.ETw[floor]) * ((s[floor] - self.sw[floor]) / (self.st[floor] - self.sw[floor])),
                                          self.ETmax])
            #self.AET_Laio = np.where(self.Rain[tim] !=0,0,self.AET_Laio) #降雨时不考虑蒸发

            ##渗漏计算【当s>sfc时发生渗漏，表现为含水率的指数形式】
            self.Lw1 = np.piecewise(s[floor],[s[floor] < self.sfc[floor],s[floor]>= self.sfc[floor]],
                                   [0,self.Ks[floor] * (np.exp(self.beta[floor] * (s[floor] - self.sfc[floor])) - 1) / (np.exp(self.beta[floor] * (1 - self.sfc[floor])) - 1)])
            self.AET.append(float(self.AET_Laio))
            self.Lw.append(float(self.Lw1))

    ##多层净含水率的计算
    def Net_s(self):
        self.Net_s0 = ((self.s_Input + self.AET[1]) - (self.Inter1 + self.AET[0] + self.Lw[0]+self.Roff))/(self.n[0] * self.Zr[0])
        self.Net_s1 = ((self.Lw[0] + self.AET[2]) - (self.AET[1] + self.Lw[1])) / (self.n[1] * self.Zr[1])
        self.Net_s2 = (self.Lw[1] - self.AET[2] - self.Lw[2]) / (self.n[2] * self.Zr[2])
        self.rhs = lambda Data: np.array([self.Net_s0,self.Net_s1,self.Net_s2])

    #解常微分方程
    def euler_forward(self, Data,dt):
        return Data + dt * self.rhs(Data)

    def improved_euler(self, Data,dt):
        yp = Data + dt * self.rhs(Data)
        return Data + 0.5 * dt * (self.rhs(Data) + self.rhs(yp))

    def Runge_Kutta4(self, Data, dt):
        k1 = dt * self.rhs(Data)
        k2 = dt * self.rhs(Data + 0.5 * k1)
        k3 = dt * self.rhs(Data + 0.5 * k2)
        k4 = dt * self.rhs(Data + k3)
        return np.array(Data + (k1 + 2.0 * (k2 + k3) + k4) / 6.0)

    def run_Laio_RK4(self):
        T, self.Sim = self.days, [] #模拟初始时刻;T:模拟总时长;Sim:模拟含水率
        s = np.array(self.s_init)   #s0:模拟初始含水率
        Cov0 = 0.25  #给定初始作物盖度
        results_avg = np.array([[0, 0, 0, 0, 0, 0, 0, 0, 0, 0]])  #初始存储占位
        for tim in tqdm(np.arange(0,T,1)):  #外部时间推进【tqdm：进度条显示】
            #TT = tim/365
            self.water_input(tim)
            if self.goal == 3: #情景分析时调用作物生长函数
                self.Crop_growth(Cov0,s[0],tim)  #作物生长表层土壤水分有关
            self.water_loss(tim, s)
            self.Net_s()
            self.Sim.append(s)
            w = np.array(s*self.n*self.Zr)  #各层储水量
            Cov = self.Cov[tim] if self.goal == 1 or 2  else self.grow_Cov
            print(s,Cov)
            ends = np.array([[tim, s.tolist(), w.tolist(),self.Rain[tim], self.AET, self.Inter1, self.Lw, self.Roff, Cov, self.A]])
            results_avg = np.concatenate((results_avg, ends))
            s = self.Runge_Kutta4(s, self.dt) #第t+1天含水率
            s = np.where(s<0,0,s)
            s = np.where(s>1,1,s)
            Cov0 = self.grow_Cov if self.goal == 3 else None
        results_avg = pd.DataFrame(results_avg,
                                   columns=['tim(day)',
                                            'Volumetric soil water content (nondim)',
                                            'Water storage(mm)',
                                            'Rain (mm/d)',
                                            'AET_imitate (mm/d)',
                                            'Inter_imitate (mm/d)',
                                            'Lw_imitate (mm/d)',
                                            'Roff_imitate (mm/d)',
                                            'Grow_Cov (%)',
                                            'Soil erosion loss (g/m2*day)'])
        results_avg = results_avg[1:]  #去除第一行占位的初始值
        
        if self.goal == 1:
            results_avg.to_csv('result/MC_result/Simulated_fenxiang/Crop{}/第{}次反演模拟.txt'.format(self.Crop_type, self.order),index=False)
            wb = openpyxl.load_workbook("result/MC_result/Simulated_s/Crop{}.xlsx".format(self.Crop_type))
            for day, S_day in enumerate(self.Sim, 1):
                for floor, name in enumerate(wb.sheetnames):
                    ws = wb[name]
                    ws.cell(row=day, column=self.order).value = S_day[floor]
                    wb.save("result/MC_result/Simulated_s/Crop{}.xlsx".format(self.Crop_type))
            wb.close()  #关闭
        if self.goal == 2:
            results_avg.to_csv('result/ME_result/Simulated_fenxiang/Crop{}/第{}次正演验证.txt'.format(self.Crop_type, self.order), index=False)
            wb = openpyxl.load_workbook("result/ME_result/Simulated_s/Crop{}.xlsx".format(self.Crop_type))
            for day, S_day in enumerate(self.Sim, 1):
                for floor, name in enumerate(wb.sheetnames):
                    ws = wb[name]
                    ws.cell(row=day, column=self.order).value = S_day[floor]
                    wb.save("result/ME_result/Simulated_s/Crop{}.xlsx".format(self.Crop_type))
            wb.close()  #关闭
       
        #返回用于敏感性分析的模拟值
        if self.goal == 4:
            fy_t = [0, 16, 31, 45, 61, 68, 81, 94, 116, 127, 137, 159, 181]
            self.Sim_s0, self.Sim_s1, self.Sim_s2 = [], [], []
            for s, index in zip(self.Sim, np.arange(len(self.Sim))):
                if index in fy_t:
                    self.Sim_s0.append(s[0]), self.Sim_s1.append(s[1]), self.Sim_s2.append(s[2])
            self.Sim_s0, self.Sim_s1, self.Sim_s2 = np.array(self.Sim_s0), np.array(self.Sim_s1), np.array(self.Sim_s2)
            Sensitivity_data = np.hstack((self.Sim_s0, self.Sim_s1, self.Sim_s2))
            return Sensitivity_data
        
        #返回用于反演与验证的实测与模拟含水率
        """
        if self.goal == 1 or 2:
            if self.goal == 1:
                fy_t = [0, 32, 43, 74, 108, 145, 166, 191] 
            else:
                fy_t = [0, 16, 31, 45, 61, 68, 81, 94, 116, 127, 137, 159, 181]
            self.Sim_s0, self.Sim_s1, self.Sim_s2 = [], [], []
            self.Obs_s0, self.Obs_s1, self.Obs_s2 = self.s_actual[0], self.s_actual[1], self.s_actual[2]
            for s, index in zip(self.Sim, np.arange(len(self.Sim))):
                if index in fy_t:
                    self.Sim_s0.append(s[0]), self.Sim_s1.append(s[1]), self.Sim_s2.append(s[2])
            self.Sim_s0, self.Sim_s1, self.Sim_s2 = np.array(self.Sim_s0), np.array(self.Sim_s1), np.array(self.Sim_s2)
            fy_data = ([self.Sim_s0, self.Obs_s0], [self.Sim_s1, self.Obs_s1], [self.Sim_s2, self.Obs_s2])
            return fy_data
       """
       
        if self.goal == 3:
            ##只考虑气候变化
            results_avg.to_csv('result/SA_result/CC/{}/Crop{}/第{}种气候变化类型.txt'.format(Model,self.Crop_type, self.order),index=False)
            ##既考虑气候变化又考虑土地利用方式转换
            #当order等于时即为不考虑气候变化
            #results_avg.to_csv('result/SA_result/CC_LUTC/{}/Crop{}_{}/植被A种植{}年后开始转换/第{}种气候变化类型.txt'.format(Model,self.Crop_typeA, self.Crop_type, self.X, self.order),index=False)   
   
    
   #生长计算模块(自然撂荒)
    def Crop_growth(self,Cov0,s0,tim):
            #r:生长速率; K:最大生长容量；L:最大生物损失量； Gs：最适合土壤含水率, Ds：死亡时对应的土壤含水率, P：形状参数(2,4.6)
            if self.Crop_type == 1:
                self.grow_Cov = self.Cov[tim]
                self.grow_Kc = self.Kc[tim]
                self.K = np.max(self.Cov)
            elif self.Crop_type == 2 or 3:  
                self.Gs, self.P = self.sfc[0]*self.n[0], 2
                self.Ds = self.st[0]*self.n[0]
                [self.K,self.r,self.L] = [1,5.546*0.001,4.196*0.001] if self.Crop_type == 2 else [1,5.479*0.001,2.632*0.001]
                self.R = self.r * Cov0 * (1 - Cov0 / self.K) * self.dt  #增加计算
                self.D_R = self.L * (s0*self.n[0] - self.Gs) ** self.P / (((s0*self.n[0] - self.Gs) ** self.P)+self.Ds**self.P)*self.dt  #水分胁迫
                self.D_R = np.where(s0 > self.sfc[0], 0, self.D_R)  #假设当含水率高于最适合植被生长时，死亡量为0
                #净生长速率
                self.grow_Cov = Cov0 + (self.R - self.D_R)
                self.grow_Cov = np.where(self.grow_Cov>0,self.grow_Cov,0)  #盖度不能小于0
                self.grow_Cov = np.where(self.grow_Cov>1,1,self.grow_Cov)  #盖度不能大于1
                self.grow_Kc = np.where(self.grow_Cov>0.1,1.59274*self.grow_Cov-0.03876,1) if self.Crop_type == 2 else np.where(self.grow_Cov > 0.1,0.94195*self.grow_Cov+0.43851, 1)#盖度小于10%当裸地处理
                """
                [self.K,gmax,dmax] = [1,0.003928571,0.008412698] if self.Crop_type == 2 else [1,0.005479452,0.002631579]  #gamx:最大生长速率；dmax：最大死亡速率
                self.gs = np.piecewise(s0,[s0 <= self.sw[0],self.sw[0] < s0 <= self.sfc[0],self.sfc[0] < s0 <= 1],[0,gmax * (s0-self.sfc[0])/(self.sfc[0]-self.sw[0]),gmax])
                self.gs = np.where(self.gs>0,self.gs,0)
                self.ds = np.piecewise(s0,[s0 <= self.sw[0],self.sw[0] < s0 <= self.sfc[0],self.sfc[0] < s0 <= 1],[dmax,dmax * (s0-self.sfc[0])/(self.sfc[0]-self.sw[0]),0])
                self.ds = np.where(self.ds>0,self.ds,0)
                self.grow_Cov = Cov0 + (self.gs - self.ds)
                self.grow_Cov = np.where(self.grow_Cov>0,self.grow_Cov,0)  #盖度不能小于0
                self.grow_Cov = np.where(self.grow_Cov>1,1,self.grow_Cov)  #盖度不能大于1
                self.grow_Kc = 1.59274*self.grow_Cov-0.03876 if self.Crop_type == 2 else 0.94195*self.grow_Cov+0.43851
                self.grow_Kc = np.where(self.grow_Kc>1.05,1.05,self.grow_Kc) if self.Crop_type == 2 else np.where(self.grow_Kc>1.2,1.2,self.grow_Kc) 
               """
            
